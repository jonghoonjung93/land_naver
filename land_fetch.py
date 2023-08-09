from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from datetime import timedelta
from module.etc_util import printL, mode_check, tele_push_admin
import telegram
import asyncio, json
import time, datetime
import csv
from openpyxl import load_workbook	# 엑셀 읽기위해 추가
import sqlite3
import sys

global_var = 0
global_err = 0
global_new = 0
global_zero = 0
global_msg_contents = []

def land_naver(building):
	flag = True
	if flag:	# 환경설정 부분
		options = Options()
		
		# global hostname
		if mode_check() == 'TEST':
			# options.add_argument("headless") #크롬창이 뜨지 않고 백그라운드로 동작됨
			pass
		else:
			options.add_argument("headless") #크롬창이 뜨지 않고 백그라운드로 동작됨

		# 아래 옵션 두줄 추가(NAS docker 에서 실행시 필요, memory 부족해서)
		options.add_argument('--no-sandbox')
		options.add_argument('--disable-dev-shm-usage')

		# headless로 동작할때 element를 못찾으면 아래 두줄 추가 (창크기가 작아서 못찾을수 있음)
		# options.add_argument("start-maximized")
		# options.add_argument("window-size=1920,1080")
		options.add_argument("window-size=1920,2160")	# 창크기가 작아서 전체 내용이 표시되지 않아서 크게 키움
		# options.add_argument("window-size=1300,2000")	# 이 크기로는 동그라미 클릭 오류가 가끔발생
		# options.add_argument("disable-gpu")
		# options.add_argument("lang=ko_KR")

		flag = False	# ----- TEST 환경에서 임시로 headless 사용할때만 True 로 놓고 사용 (아직 잘안됨, 해결필요) ---------------
		if flag:
			options.add_argument('headless')

		# 아래는 headless 크롤링을 막아놓은곳에 필요 (user agent에 HeadlessChrome 이 표시되는걸 방지)
		options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36')
			
		# config.json 파일처리 ----------------
		with open('config.json','r') as f:
			config = json.load(f)
		url = config[building]['URL']
		naver_bld_id = config[building]['NAVER_BLD_ID']
		address = config[building]['ADDRESS']
		memo = config[building]['MEMO']
		# ------------------------------------

		driver = webdriver.Chrome(options=options)

    # 건물정보 csv 파일처리 ----------------
	flag = False	# 엑셀방식으로 전환했음. 지금은 사용안함
	if flag:
		csv_data = []
		address_short = address.replace("고양시 일산동구 ","")
		file_name = f"csv/{building[3:]}_{memo}({address_short}).csv"
		with open(file_name,'r') as file:	# csv_data 라는 변수에 건물정보를 저장
			reader = csv.reader(file)
			for line in reader:
				csv_data.append(line)
		# for line in csv_data:	# line[0]:호수, line[1]:총면적, line[2]:전용면적, line[3]:세부용도
		# 	print(line[1])
	# ------------------------------------

	# 건물정보 excel 파일을 읽어서 list 변수에 넣기 ----------------------------
	flag = True
	if flag:
		xlsx_data = []
		address_short = address.replace("고양시 일산동구 ","")
		file_name = f"xlsx/{building[3:]}_{memo}({address_short}).xlsx"
		wb = load_workbook(file_name) # sample.xlsx 파일에서 wb 을 불러옴
		ws = wb.active # 활성화된 Sheet

		for x in range(1, ws.max_row + 1):
			cell1 = ws.cell(row=x, column=1).value
			cell2 = ws.cell(row=x, column=2).value
			cell3 = ws.cell(row=x, column=3).value
			if cell3 != None:	# 층 컬럼에 값이 있을때
				cell3_convert = cell3.replace("층","").replace("F","")	# 층 컬럼에
			else:	# 층 컬럼에 값이 없을때 (값이 없을때 replace 함수를 쓰면 Nontype 에러남)
				cell3_convert = cell3
			cell4 = ws.cell(row=x, column=4).value
			cell5 = ws.cell(row=x, column=5).value
			cell6 = ws.cell(row=x, column=6).value
			# cell7 = ws.cell(row=x, column=7).value
			cell8 = ws.cell(row=x, column=8).value
			cell9 = ws.cell(row=x, column=9).value
			# cell10 = ws.cell(row=x, column=10).value
			#             호수,  총면적,  전용면적, 용도,  건물명,  구분,   층 ,   전용(평)
			temp_list = [cell4, cell5, cell6, cell9, cell1, cell2, cell3_convert, cell8]
			xlsx_data.append(temp_list)
	# ------------------------------------------------------------------


	# 페이지 접속
	driver.get(url)
	time.sleep(2)	# 20230718 : 2->4 변경

	new_list = []	# 신규 매물 리스트 (알림 대상)

	# 해당건물 클릭
	try:
		driver.find_element(By.ID, naver_bld_id).find_element(By.CLASS_NAME, 'marker_transparent').click()	# 건물 동그라미 클릭
	except:
		printL(f"{building}({memo}) marker click retry...")
		time.sleep(2)
		try:
			driver.get(url)
			time.sleep(3)
			driver.find_element(By.ID, naver_bld_id).find_element(By.CLASS_NAME, 'marker_transparent').click()	# 건물 동그라미 클릭2
		except:
			time.sleep(1)
			printL(f"ERROR : {building}({memo}) marker click failed... return")
			html_content = driver.page_source
			if naver_bld_id in html_content:	# 클릭을 실패했는데 건물ID가 html 에 있는 경우 (진짜 클릭실패한 경우)
				printL(f"naver_bld_id({naver_bld_id})가 html 안에 존재함")
				driver.quit()
				raise RuntimeError(f"RuntimeError :{building}({memo}) 건물 동그라미 클릭 실패")	# 에러 return
			else:								# 클릭을 실패했는데 건물ID가 없는 경우 (해당 건물에 매물이 없어서 클릭을 못한 경우)
				printL(f"naver_bld_id({naver_bld_id})가 html 안에 존재하지 않음, 해당 건물에 매물이 한개도 없음")
				global global_zero
				global_zero = global_zero + 1
				driver.quit()
				return(new_list)	# 정상 return
	# time.sleep(1000)

	time.sleep(2)

	scroll_bar = driver.find_element(By.XPATH, '//*[@id="listContents1"]/div')	# 좌측 매물리스트 스크롤바
	
	# for _ in range(10):	# 스크롤바 내리기 10회 반복
	while True:	# 스크롤바 내리기 무한반복 (필요한만큼)
		# print(driver.execute_script("return document.documentElement.scrollHeight"))
		new_height = driver.execute_script("return arguments[0].scrollHeight", scroll_bar)
		# print(new_height)
		driver.execute_script("arguments[0].scrollBy(0,2000);", scroll_bar)	# 스크롤바 내리기
		time.sleep(0.5)
		driver.execute_script("arguments[0].scrollBy(0,2000);", scroll_bar)	# 스크롤바 내리기
		time.sleep(0.5)
		last_height = driver.execute_script("return arguments[0].scrollHeight", scroll_bar)
		# print(last_height)
		if new_height == last_height:	# 더이상 내려가지 않을때
			break

	# 매물 가져오기
	items = driver.find_element(By.CLASS_NAME, 'item_list.item_list--article').find_elements(By.CLASS_NAME, 'item.false')
	# items_len = len(driver.find_element(By.CLASS_NAME, 'item_list.item_list--article').find_elements(By.CLASS_NAME, 'item.false'))
	printL(f"building :  {building}({memo}), {len(items)}")	# 매물 개수 확인
	# printL(f"building :  {building}({memo}), {items_len}")	# 매물 개수 확인

	# 날짜 (오늘)
	current_time = datetime.datetime.now()
	formatted_date = current_time.strftime("%Y%m%d")
	# 어제,그제 날짜 yyyymmdd 로 만들기
	today = current_time.date()
	yesterday_form = today - timedelta(days=1)
	yesterday = yesterday_form.strftime("%Y%m%d")
	# delete_day_form = today - timedelta(days=2)	# 삭제대상, 2일전 데이터는 삭제처리
	delete_day_form = today - timedelta(days=3)	# 삭제대상, 3일전 데이터는 삭제처리
	delete_day = delete_day_form.strftime("%Y%m%d")

	# DB 접속
	conn = sqlite3.connect('land_naver.sqlite3')
	cursor = conn.cursor()

	data_list = []	# DB에 insert 대상 매물 리스트
	
	# for item in items:
	# for index in range(items_len):
	for item in items:
		# printL(index)
		# item = driver.find_element(By.CLASS_NAME, 'item_list.item_list--article').find_elements(By.CLASS_NAME, 'item.false')[index]
		name = item.find_element(By.CLASS_NAME, 'text').text
		# printL(name)
		# --- 매물번호 가져오기 (item_number) 매물을 하나하나 클릭해서 변경된 URL에서 값을 가져옴 ------------------------- #
		try:	# "네이버에서 보기" 버튼이 있는 경우
			naver_view = item.find_element(By.CLASS_NAME, 'label.label--cp')
			# printL(f"naver_view : {naver_view.text}")
			if naver_view.text == "네이버에서 보기":
				naver_view.click()
		except:	# "네이버에서 보기" 버튼이 없는 경우
			try:
				item.find_element(By.CLASS_NAME, 'text').click()
			except:
				pass
		# wait = WebDriverWait(driver, 10)
		# new_url = wait.until(EC.url_changes(driver.current_url))
		try:
			new_url = driver.current_url
			# printL(f"Changed URL: {new_url}")
			split_url = new_url.split("=")
			last_part = split_url[-1]
			item_number = last_part.split("&")[0]
			# printL(item_number)
			naver_bld_id = item_number	# naver_bld_id 는 클릭하고 나면 필요없으니, 여기부터는 이걸 매물번호로 활용
		except:
			naver_bld_id = "0000000000"
		# ------------------------------------------------------------------------------------------------ #
		type = item.find_element(By.CLASS_NAME, 'type').text
		price = item.find_element(By.CLASS_NAME, 'price').text
		info_area_type = item.find_element(By.CLASS_NAME, 'info_area').find_element(By.CLASS_NAME, 'type').text
		info_area_spec = item.find_element(By.CLASS_NAME, 'info_area').find_element(By.CLASS_NAME, 'spec').text
		agent_name_list = item.find_elements(By.CLASS_NAME, 'agent_name')
		agent_name = agent_name_list[1].text
		info_count = info_area_spec.count(",")
		if info_count == 2:
			var1, var2, var3 = info_area_spec.split(',')
		elif info_count == 1:
			var1, var2 = info_area_spec.split(',')
			var3 = ""
		else:
			var1 = info_area_spec
			var2 = ""
			var3 = ""
		size_total, size_real = var1.replace('m²','').split('/')	# 총면적, 전용면적 가져오기
		# floor = var2.replace(' ','')[0]
		try:
			floor, etc99 = var2.replace(' ','').split('/')
		except:
			floor = var2.replace(' ','')[0]
		ho = ''
		try:	# csv 파일이 헤더만 있는경우 line[1]을 못가져오고 에러나는것에 대한 처리
			# for line in csv_data[1:]:	# line[0]:호수, line[1]:총면적, line[2]:전용면적, line[3]:세부용도
			for line in xlsx_data[1:]:	# line[0]:호수, line[1]:총면적, line[2]:전용면적, line[3]:세부용도, 4:건물명, 5:구분, 6:층, 7:전용(평)
				csv_size1 = int(float(line[1]))	# 현황이 없으면 이부분에서 에러남
				# if csv_size1 == int(size_total) and floor == line[0][0]:	# 총면적이 같고, 층수가 같으면 -> 호수의 앞자리만 봤더니 1층 찾는데 10층까지 딸려옴
				if csv_size1 == int(size_total) and floor == line[6]:	# 총면적이 같고, 층수가 같으면 (엑셀의 별도 호수 컬럼 이용)
					# print(f"size1={size1}:{floor}, csv_size1={csv_size1}:{line[1]}:{line[0]}")
					ho = ho + line[0]	# 같은 호수가 여러개일수 있음
					if ho is not None:	# 같은 호수가 여러개면 ,를 이용해서 나열
						ho = ho + ","
					# size_real = line[7]
			if ho.endswith(","):	# 제일 끝에 , 가 있으면 삭제
				ho = ho[:-1]
		except:
			ho = '현황없음'
		# 중소형사무실 | 월세 | 3,000/190 | 사무실 | 251/135m², 2/7층, 북서향 | 251 | 135 | 2 |  하임빌공인중개사무소 | 214,221
		# print(f"{name} | {type} | {price} | {info_area_type} | {info_area_spec} | {size_total} | {size_real} | {floor} |  {agent_name} | {ho}")
		bld_id = building
		try:
			price_base, price_mon = price.split("/")
		except:
			price_base = price
			price_mon = ""
		
		# 해당 매물이 기존에 DB에 있는지 확인 (어제자와 비교)
		select_sql = f"SELECT * FROM land_item WHERE date='{yesterday}' AND bld_id = '{bld_id}' AND price = '{price}' AND info_area_spec = '{info_area_spec}' AND agent_name = '{agent_name}'"
		cursor.execute(select_sql)
		result_select = cursor.fetchone()
		# print(result_select)

		if result_select is not None:	# 매물을 어제자 DB에서 조회했는데 이미 있던 매물일 경우
			# print("있다")
			new = ""
			tuple1 = (formatted_date, bld_id, memo, address_short, url, naver_bld_id, name, type, price, price_base, price_mon, info_area_type, info_area_spec, size_total, size_real, floor, agent_name, ho, new)
		else:	# 매물이 어제자 DB에 없는 새로운 매물일 경우
			# print("없다. 새로운거 발견!!!!")
			new = "O"
			tuple1 = (formatted_date, bld_id, memo, address_short, url, naver_bld_id, name, type, price, price_base, price_mon, info_area_type, info_area_spec, size_total, size_real, floor, agent_name, ho, new)
			new_list.append(tuple1)
			global global_new
			global_new = global_new + 1

		data_list.append(tuple1)	# tuple 을 list 에 추가
		result = f"{name} | {type} | {price} | {info_area_type} | {info_area_spec} | {agent_name} | {ho}"

	# print(new_list)

    # DB insert 처리
	delete_sql1 = f"DELETE FROM land_item WHERE date = '{delete_day}' AND bld_id = '{bld_id}'"	# 그저께 데이터 삭제
	cursor.execute(delete_sql1)
	delete_sql2 = f"DELETE FROM land_item WHERE date = '{formatted_date}' AND bld_id = '{bld_id}'"	# 당일자 삭제 (이미 있을때 다시 넣기 위해서)
	cursor.execute(delete_sql2)
	cursor.executemany("INSERT INTO land_item (date, bld_id, memo, address, url, naver_bld_id, name, type, price, price_base, price_mon, info_area_type, info_area_spec, size_total, size_real, floor, agent_name, ho, new) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);", data_list)
	conn.commit()
	conn.close()

	# time.sleep(1000)
	driver.quit()
	return(new_list)

# telegram 메세지 발송함수
async def tele_push(content): #텔레그램 발송용 함수
	if mode_check() == 'TEST':
		telegram_target = 'TELEGRAM-TEST'	# TEST 모드
	else:
		telegram_target = 'TELEGRAM'	# ONLINE 모드
	# telegram_target = 'TELEGRAM'	# 강제 ONLINE 모드
	# config.json 파일처리 ----------------
	with open('config.json','r') as f:
		config = json.load(f)
	token = config[telegram_target]['TOKEN']
	chat_ids = config[telegram_target]['CHAT-ID']
	# ------------------------------------

	current_time = datetime.datetime.now()
	# formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")

	result = [0, 0, 0, 0]	# 대상자, 성공, 재시도, 실패 건수
	result[0] = len(chat_ids)
	bot = telegram.Bot(token = token)

	# 실패시 재시도를 여기서 하도록 변경
	for chat_id in chat_ids:
		# printL(f"chat_id : {chat_id}")
		send_retry = 0
		while True:	# 텔레그램 발송이 혹시 실패하면 최대 3회까지 성공할때까지 재시도
			try:
				# await bot.send_message(chat_id, formatted_time + "\n" + content, parse_mode = 'Markdown', disable_web_page_preview=True)
				await bot.send_message(chat_id, content, parse_mode = 'Markdown', disable_web_page_preview=True)
				printL(f"-- SEND success!!! : {chat_id}")
				result[1] = result[1] + 1	# 성공
				break
			except:
				send_retry = send_retry + 1
				result[2] = result[2] + 1	# 재시도
				printL(f"-- tele_push failed!!! ({send_retry}) : chat_id = {chat_id}")
				time.sleep(3)
				if send_retry == 3:
					printL(f"-- tele_push aborted!!! : chat_id = {chat_id}")
					result[3] = result[3] + 1	# 실패
					printL(f"-- content : {content}")
					break
			else:	# 정상작동시
				pass
			finally:	# 마지막에 (정상,에러 상관없이)
				pass
	printL("-------------------------------------")
	return(result)

	global global_var
	global_var = global_var + 1	# 보낸 메세지가 있으면 +1씩 올라감

def send_lists(lands):	# 전일자와 비교해서 현재 새로운 매물리스트를 메세지로 발송
	for land in lands:
		try:
			lists = land_naver(land)
		except KeyboardInterrupt:
			printL("KeyboardInterrupt 발생. 프로그램 종료합니다.")
			sys.exit(1)
		except RuntimeError as e:	# 동그라미 클릭 에러시
			global global_err
			global_err = global_err + 1
			printL(f"{land} : ERROR! land_naver 에러 ({e}). count({global_err})")
			continue
		except Exception as e:
			# global global_err
			global_err = global_err + 1
			printL(f"{land} : ERROR! land_naver 알수없는 에러. count({global_err}), err({e})")
			continue
		else:	# 정상완료시
			# print(len(lists))
			if len(lists) == 0:	# 어제자와 비교했을때 추가된 건이 없을때
				# print("추가 매물이 없음")
				pass
			else:
				# print("lists is not None")
				msg_content = ""
				for list in lists:
					if list[17] == "":
						ho = "Not found"
					else:
						ho = list[17]
					# if msg_content is not None:
					# 	msg_content = msg_content + "\n"
					# tuple1 = (formatted_date, bld_id, memo, address_short, url, naver_bld_id, name, type, price(8), price_base(9), price_mon(10), info_area_type(11), info_area_spec(12), size_total(13), size_real(14), floor, agent_name, ho)
					deep_link = f"https://new.land.naver.com/offices?articleNo={list[5]}"
					num = round(float(list[14])*0.3025, 2)
					pyoung = f"m²({num}평)\n -"
					list_12 = list[12].replace("m²,",pyoung)
					# msg_content = f"{msg_content}\*{list[2]}({list[3]}): {list[6]} {list[7]}\n - {list[8]} {list[12]}\n - {list[16]}\n - 호수 : _{ho}_\n"
					msg_content = f"{msg_content}\*[{list[2]}({list[3]})]({deep_link}): {list[6]} {list[7]}\n - {list[8]} {list_12}\n - {list[16]}\n - 호수 : _{ho}_\n"
				printL(msg_content)
				global global_msg_contents
				global_msg_contents.append(msg_content)		# 메세지를 바로 보내지 않고 global list 변수에 담아놓기
				# asyncio.run(tele_push(msg_content)) #텔레그램 발송 (asyncio를 이용해야 함)
		finally:	# 정상,완료 상관없이 마지막에
			pass

def initial_lists(lists):	# 건물 처음으로 추가시에 메세지는 보내지 않고 오늘 매물만 DB에 insert 하고 싶을때 사용
	msg_content = ""
	for list in lists:
		if list[17] == "":
			ho = "Not found"
		else:
			ho = list[17]
		# if msg_content is not None:
		# 	msg_content = msg_content + "\n"
		msg_content = f"{msg_content}\*{list[2]}({list[3]}): {list[6]} {list[7]}\n - {list[8]} {list[12]}\n - {list[16]}\n - 호수 : _{ho}_\n"
	printL(msg_content)
	printL(f"COUNT, {len(lists)}")
	global global_var
	global_var = global_var + 1	# 초기화 작업시 "새로운 매물이 없음" 메세지 보내지는거 방지. (뭔가 이미 보낸것처럼 +1)

# lists = land_naver('BLD2-19')
# send_lists(lists)

# ----------- MAIN ------------- 

#--- Start Time 기록 -----
flag = True
if flag:
	current_time = datetime.datetime.now()
	formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
	start_time = current_time.strptime(formatted_time, "%Y-%m-%d %H:%M:%S")
	printL(f"-- START : {start_time}")
	printL(f"-- {mode_check()} MODE")

# initial_lists(land_naver('BLD2-18'))	# 초기화할때 (처음에 건물 추가할때는 이렇게 넣어야 됨)
# send_lists(land_naver('BLD2-18'))	# 변경전 예전방식
#------- 각 건물별로 실행 ----------------------
flag = True
if flag:
	lands_list = [
		'BLD1-01',
		'BLD1-02',
		'BLD1-03',
		'BLD1-04',
		'BLD1-05',	#1차 동하넥서스
		'BLD1-06',	#1차 성우사카르
		'BLD1-07',	#1차
		'BLD1-08',
		'BLD1-09',	#1차
		'BLD1-10',
		'BLD1-11',	#1차
		'BLD1-12',	#1차 77건짜리 동그라미 선택 실패 사례발생, 화면크기 키워서 해결
		'BLD1-13',
		'BLD1-14',
		'BLD1-17',
		'BLD1-18',
		'BLD1-19',
		'BLD1-20',
		'BLD1-21',
		'BLD1-23',
		'BLD1-24',
		'BLD1-25',
		'BLD1-26',
		'BLD1-27',
		'BLD1-28',
		'BLD1-29',
		'BLD1-30',
		'BLD1-31',
		'BLD1-32',
		'BLD1-33',
		'BLD1-34',
		'BLD1-36',
		'BLD2-01',
		'BLD2-02',	#1차
		'BLD2-03',	#1차
		'BLD2-04',	#1차
		'BLD2-06',
		'BLD2-07',
		'BLD2-08',
		'BLD2-09',	#1차
		'BLD2-11',	#1차
		'BLD2-12',	#1차 원장내용없는 사태 발생, 에러처리완료
		'BLD2-13',	#1차
		'BLD2-14',	#1차
		'BLD2-15',
		'BLD2-16',	#1차 하임빌
		'BLD2-17',
		'BLD2-18',	#1차 로데오탑
		'BLD2-19',	#1차
		'BLD2-20',	#1차
		'BLD2-21',	#1차
		'BLD2-22',	#1차
		'BLD2-24',
		'BLD2-27',
		'BLD2-28',
		'BLD2-29'
	]

# lands_list = ['BLD2-27','BLD2-24']	# 한개씩 실행할때
send_lists(lands_list)

# printL(f"global_msg_contents({len(global_msg_contents)}) : {global_msg_contents}")

flag_sms = True	# 메세지 발송여부 (관리자 메세지 포함)
#--- 모아뒀던 메세지 발송 ----- 사용안함
flag = False
if flag:
	tele_result_sum = [0, 0, 0, 0]
	# if global_var == 0:	# 전체적으로 보낼 메세지가 1건도 없을때
	if len(global_msg_contents) == 0:	# 보낼 메세지 list 변수안에 1건도 없을때
		if global_var == 0:		# 초기화 작업만 했을때는 메세지는 없지만 global_var는 0 이 아님
			msg_content = " - 새로운 매물이 없음"
			printL(msg_content)
			if flag_sms:
				tele_result = asyncio.run(tele_push(msg_content)) #텔레그램 발송 (asyncio를 이용해야 함)
				tele_result_sum[0] = tele_result[0]	#발송 대상자
				for i in range(1,4):	# 발송상태 (정상,재시도,실패)건수
					tele_result_sum[i] = tele_result_sum[i] + tele_result[i]
			printL(f"--- telegram 메세지 발송 결과(성공,실패) : {tele_result_sum}")
	else:	# 보낼 메세지가 있을때 (list 에 있는 내용 다 보냄)
		printL(f"global_msg_contents ({len(global_msg_contents)})개")
		for msg in global_msg_contents:
			if flag_sms:
				tele_result = asyncio.run(tele_push(msg)) #텔레그램 발송 (asyncio를 이용해야 함)
				tele_result_sum[0] = tele_result[0]	#발송 대상자
				for i in range(1,4):	# 발송상태 (정상,재시도,실패)건수
					tele_result_sum[i] = tele_result_sum[i] + tele_result[i]
		printL(f"--- telegram 메세지 발송 결과(성공,실패) : {tele_result_sum}")

#--- DB 접속 및 당일발송 대상 초기화 -----------
flag = True
if flag:
	conn = sqlite3.connect('land_naver.sqlite3')
	cursor = conn.cursor()

	# 날짜 (오늘)
	current_time = datetime.datetime.now()
	formatted_date = current_time.strftime("%Y%m%d")
	# 어제,그제 날짜 yyyymmdd 로 만들기
	today = current_time.date()
	delete_day_form = today - timedelta(days=5)	# 삭제대상, 5일전 데이터는 삭제처리
	delete_day = delete_day_form.strftime("%Y%m%d")

	delete_day_sql = f"DELETE FROM message_list WHERE date = '{delete_day}'"	# 삭제대상(과거) 데이터 삭제
	cursor.execute(delete_day_sql)
	today_delete_message_list_sql = f"DELETE FROM message_list WHERE date = '{formatted_date}' AND send_yn = '0'"	# 당일자 삭제 (이미 있을때 다시 넣기 위해서)
	cursor.execute(today_delete_message_list_sql)

#--- DB 저장함수 ----
def db_insert(msg):
	if mode_check() == 'TEST':
		telegram_target = 'TELEGRAM-TEST'	# TEST 모드
	else:
		telegram_target = 'TELEGRAM'	# ONLINE 모드
	# telegram_target = 'TELEGRAM'	# 강제 ONLINE 모드
	# config.json 파일처리 ----------------
	with open('config.json','r') as f:
		config = json.load(f)
	token = config[telegram_target]['TOKEN']
	chat_ids = config[telegram_target]['CHAT-ID']
	# ------------------------------------
	
	db_result = [0, 0, 0, 0]
	insert_list = []
	for chat_id in chat_ids:
		# insert
		# id 컬럼 = sequence 컬럼
		time = "0900"
		datetime_1 = ""
		job = "land_fetch.py"
		send_yn = "0"	# 0: 발송대기, 1:발송완료
		retry = "0"
		result = ""
		channel = "TELEGRAM"
		chat_room = token
		# chat_id = ""
		message = msg
		parse_mode = "Markdown"
		preview = "Yes"
		etc1 = etc2 = etc3 = etc4 = ""
		row_tuple = (None, formatted_date, time, datetime_1, job, send_yn, retry, result, channel, chat_room, chat_id, message, parse_mode, preview, etc1, etc2, etc3, etc4)
		insert_list.append(row_tuple)
	try:
		# db insert
		cursor.executemany("INSERT INTO message_list (id, date, time, datetime, job, send_yn, retry, result, channel, chat_room, chat_id, message, parse_mode, preview, etc1, etc2, etc3, etc4) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);", insert_list)
		db_result[0] = len(chat_ids)	# 대상자수
		db_result[1] = db_result[1] + len(insert_list)	# 대상건수
	except:
		# insert fail
		db_result[3] = db_result[3] + len(insert_list)	# 실패건수

	return(db_result)

#--- 모아뒀던 메세지 DB에 저장 ----
flag = True
if flag:
	tele_result_sum = [0, 0, 0, 0] # 대상자count, 정상count, 재시도count, 실패count
	# if global_var == 0:	# 전체적으로 보낼 메세지가 1건도 없을때
	if len(global_msg_contents) == 0:	# 보낼 메세지 list 변수안에 1건도 없을때
		if global_var == 0:		# 초기화 작업만 했을때는 메세지는 없지만 global_var는 0 이 아님
			msg_content = " - 새로운 매물이 없음"
			printL(msg_content)
			if flag_sms:
				# tele_result = asyncio.run(tele_push(msg_content)) #텔레그램 발송 (asyncio를 이용해야 함)
				tele_result = db_insert(msg_content) #텔레그램 발송 (asyncio를 이용해야 함)
				tele_result_sum[0] = tele_result[0]	#발송 대상자
				for i in range(1,4):	# 발송상태 (정상,재시도,실패)건수
					tele_result_sum[i] = tele_result_sum[i] + tele_result[i]
			printL(f"--- telegram 발송용 DB insert 결과(성공,실패) : {tele_result_sum}")
	else:	# 보낼 메세지가 있을때 (list 에 있는 내용 다 보냄)
		printL(f"global_msg_contents ({len(global_msg_contents)})개")
		for msg in global_msg_contents:
			if flag_sms:
				# tele_result = asyncio.run(tele_push(msg)) #텔레그램 발송 (asyncio를 이용해야 함)
				tele_result = db_insert(msg) #DB insert 함수 실행
				tele_result_sum[0] = tele_result[0]	#발송 대상자
				for i in range(1,4):	# 발송상태 (정상,재시도,실패)건수
					tele_result_sum[i] = tele_result_sum[i] + tele_result[i]
		printL(f"--- telegram 발송용 DB insert 결과(성공,실패) : {tele_result_sum}")

#--- End Time 기록 -----
flag = True
if flag:
	current_time = datetime.datetime.now()
	formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
	end_time = current_time.strptime(formatted_time, "%Y-%m-%d %H:%M:%S")
	elap_time = end_time - start_time
	printL(f"-- END : {end_time}")
	printL(f"-- Elapsed time : {elap_time}")

#--- 관리자 결과 보고 -----
flag = True
if flag:
	# DB에 발송대기중인 건수 몇건인지 확인
	count_message_list_sql = f"SELECT count(*) FROM message_list WHERE date='{formatted_date}' AND send_yn = '0'"
	cursor.execute(count_message_list_sql)
	row_count = cursor.fetchone()[0]

	# completed_message = f"\[완료] 총건물 : {len(lands_list)}개\n - 추가된 매물/건물 : {global_new}개 / {len(global_msg_contents)}개\n - 에러처리된 건물 : {global_err}개\n - 매물이 없는 건물 : {global_zero}개\n - Elapsed time : {elap_time}"
	completed_message = (
		f"\[fetch 완료] 총건물 : {len(lands_list)}개\n"
		f" - 추가된 매물/건물 : {global_new}개 / {len(global_msg_contents)}개\n"
		f" - 에러처리된 건물 : {global_err}개\n"
		f" - 매물이 없는 건물 : {global_zero}개\n"
		f" - 발송대기 : {tele_result_sum[0]}명, {tele_result_sum[1]}건\n"
		f" - DB확인 : {row_count} rows\n"
		f" - Elapsed time : {elap_time}"
	)
	printL(completed_message)
	if flag_sms:
		asyncio.run(tele_push_admin(completed_message)) #텔레그램 발송 (asyncio를 이용해야 함)

#--- DB close ------------
flag = True
if flag:
	conn.commit()
	conn.close()